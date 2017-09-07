#download all design examples
#Dustin Henderson

import urllib
import urllib2
import re

from openpyxl import Workbook
from downloadParFile import downloadParFile

def findDownloadLink(html):
	start = html.find(downloadIndicator)
	end = 0
	k = 1
	while(html[start + k] != "\""):
		k = k + 1
	start = start + k
	#print "guess at download url: ", html[start:start+32]
	k = 1
	while(html[start + k] != "\""):
		k = k + 1
	end = start + k
	return html[start+1:end]
	#print "url found: ", html[start+1:end], " start: ", start, " end: ", end
	
	
def findMaxPage(html):
	#print html
	returnNum = 0
	for m in re.finditer("page=", html): #note finditer cannot search for special charicters (? messes it up)
		stop = False
		j = 0
		pageNumber = ""
		while(stop == False):
			if(html[m.end() + j] == "\""):
				stop = True
			else:
				pageNumber += html[m.end() + j]
			j = j + 1
		if(returnNum < int(pageNumber)):
			returnNum = int(pageNumber)
	return returnNum

def finder(html):
	listOfUrls = []
	for m in re.finditer("href=\"/devstore/platform/", html):
		stop = False
		j = 0
		fromAltera = ""
		while(stop == False):
			if(html[m.end() + j] == "\""):
				stop = True
				print "found: ", fromAltera
				print "beginning: ", m.end(), " j : ", j
			else:
				fromAltera += html[m.end() + j]
			j = j + 1
		if(j != 1):
			print "Append j: ", j
			listOfUrls.append(urlHead + fromAltera)
	print listOfUrls
	return listOfUrls

'''*************************************************************************************************	
*************************************** Main Part Of Script ****************************************	
*************************************************************************************************'''
page = 1		#start the page at page 1 on the website
pageMax = 1	#The number of pages on the design store
referancesFound = 0
downloadsFound = 0
urlHead = 'https://cloud.altera.com/devstore/platform/'
fullUrlList = []
downloadIndicator = "Installation Package"

responce = urllib2.urlopen('https://cloud.altera.com/devstore/platform/')
html = responce.read()
pageMax = findMaxPage(html)

print "Page Max Found: ", pageMax
print "***Getting Links From Pages***"
	
while(page <= pageMax):
	responce = urllib2.urlopen( 'https://cloud.altera.com/devstore/platform/?page=' + str(page))
	html = responce.read()
	print "*************************************************************************************\n"
	print "Page: ", page 
	#search loop. finds all the instances of the search string in the HTML of the page
	fullUrlList = fullUrlList + (finder(html))
	#referancesFound = referancesFound + 1
	#print html
	print "*************************************************************************************\n"
	
	'''***** For Testing *****'''
	if(page > 2):
		break
	'''***********************'''
	page = page + 1

noLinkList = []

print "***Determine if external downlaod***"
for line in fullUrlList:
	print line
	#time.sleep(10) #for testing
	responce = urllib2.urlopen(line)
	html = responce.read()
	if html.find(downloadIndicator) != -1:
		print "Download Found"
		downloadsFound = downloadsFound + 1
		downloadParFile("https://cloud.altera.com" + findDownloadLink(html), "test" + str(downloadsFound)) #https://cloud.altera.com/devstore/platform/755/download/
	else:
		noLinkList.append(line)
		#ws['A' + str(noLinkCounter)] = "=HYPERLINK(" + line + ")"
		#noLinkCounter = noLinkCounter + 1
	#print html
	#time.sleep(10) #for testing
	referancesFound = referancesFound + 1
	
	
	'''***** For Testing *****'''
	if(referancesFound > 9):
		break
	'''***********************'''

noLinkCounter = 1
wb = Workbook()
ws = wb.active

print "*** Write to File ***"

for url in noLinkList:
	print 'A' + str(noLinkCounter)
	print "=HYPERLINK(\"" + url + "\")"
	ws['A' + str(noLinkCounter)] = "=HYPERLINK(\"" + url + "\")"
	noLinkCounter = noLinkCounter + 1
	
wb.save("noDownloadLink.xlsx")
print "\n total links found: ", referancesFound
print "\n total downloads found: ", downloadsFound
print "\n total nonDownload links found: ", noLinkCounter -1



#********************* important **************************
#******************* Do Not Delete ************************
#email tiu, jonas for upload ideas
#********************* important **************************
#******************* Do Not Delete ************************