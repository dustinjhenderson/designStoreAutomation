#Design Store Analytics
#Daniel Nielsen
#New User Experience Group
#Tabs used for indentation

# Importing packages
from __future__ import division
try :
	import openpyxl
	from openpyxl.styles import Font
except ImportError :
	print 'ERROR : Run the batch file again \nOR \nMake sure \'openpyxl\' package is installed correctly, to check enter in command prompt: \n<python -m pip install openpyxl> and re-run the script'
	exit()	
try :
	import glob
except ImportError :
	print 'ERROR : Run the batch file again \nOR \nMake sure \'glob\' package is installed correctly, to check enter in command prompt: \n<python -m pip install glob>'
	exit()	
try :
	import os
except ImportError :
	print 'ERROR : Run the batch file again \nOR \nMake sure \'os\' package is installed correctly, to check enter in command prompt: \n<python -m pip install os>'
	exit()	
try :
	import requests
except ImportError :
	print 'ERROR : Run the batch file again \nOR \nMake sure \'requests\' package is installed correctly, to check enter in command prompt: \n<python -m pip install requests>'
	exit()

#Opening workbook
print 'Executing Script'
#Error checking - making sure User followed README
def runAnalytics():
	cnt = 0
	for file in glob.glob("*.xlsx") :
		in_file = glob.glob("*.xlsx")
		cnt += 1
	if cnt == 0 :
		try :
			in_file = glob.glob("*.xls")
			if not in_file :
				print 'ERROR : No input file - please download Excel report from Design Store and place the script and Excel in the same directory'
				exit()
			wb = openpyxl.load_workbook(in_file[0])
		except openpyxl.utils.exceptions.InvalidFileException :
			print 'ERROR : .xls format not supported, please convert .xls file to .xlsx format and re-run the script'
			exit()
	elif cnt > 1 :
		print 'ERROR : \n1) Please make sure that the excel to be analyzed is the only excel document in the directory with the script an re-run the script\n2) If it is the only excel, make sure it is not open'
		exit()
		
	wb = openpyxl.load_workbook(in_file[0])
	print 'Excel File Read'
	try : 
		sheet = wb.worksheets[0]
	except KeyError:
		print 'ERROR : Missing Analytics Data Sheet!'
		exit()

	print 'Sorting Data To Respective Worksheets'	
	#Create dictionaries/tuples in which to store data
	categoryData = {}
	categoryList = []
	designData = {}
	designList = []
	familyData = {}	
	familyList = []
	devkitData = {}
	devkitList = []

	#Store data from original worksheet in the corresponding dictionary	
	for row in range(2, sheet.max_row + 1):
		category = sheet['A' + str(row)].value
		downloads = sheet['N' + str(row)].value
		categoryData.setdefault(category, 0)
		categoryData[category] += int(downloads)
		
	for row in range(2, sheet.max_row + 1):
		design = sheet['B' + str(row)].value
		downloads = sheet['N' + str(row)].value
		designData.setdefault(design, 0)
		designData[design] += int(downloads)

	for row in range(2, sheet.max_row + 1):
		family = sheet['F' + str(row)].value
		downloads = sheet['N' + str(row)].value
		familyData.setdefault(family, 0)
		familyData[family] += int(downloads)
		
	for row in range(2, sheet.max_row + 1):
		devkit = sheet['L' + str(row)].value
		downloads = sheet['N' + str(row)].value
		devkitData.setdefault(devkit, 0)
		devkitData[devkit] += int(downloads)


	#Convert dictionaries to tuples
	for category in categoryData:
		categoryList.append((category, categoryData[category]))

	for design in designData:
		designList.append((design, designData[design]))

	for family in familyData:
		familyList.append((family, familyData[family]))
		
	for devkit in devkitData:
		devkitList.append((devkit, devkitData[devkit]))

	#Sort category data based upon total downloads and store in new worksheet	
	categoryList = sorted(categoryList, key=lambda view: view[1], reverse=True)
	try :
		sheet = wb.get_sheet_by_name('DownloadsByCategory')
	except KeyError:
		wb.create_sheet(title = 'DownloadsByCategory')
		sheet = wb.get_sheet_by_name('DownloadsByCategory')
	fontobj = Font(bold=True)
	sheet['A1'] = 'Category'
	sheet['A1'].font = fontobj
	sheet['B1'] = 'Total Downloads'
	sheet['B1'].font = fontobj
	sheet['C1'] = 'Description'
	sheet['C1'].font = fontobj

	i = 2
	for i in range(len(categoryList)) :
		sheet['A'+str(i+2)] = categoryList[i][0] 
		sheet['B'+str(i+2)] = categoryList[i][1]

	#Write in category descriptions
	DE = 'Design Example'
	T = 'Tutorial'
	DEODS = 'Design Example \ Outside Design Store'
	DELODS = 'Design Example \ Linux Outside Design Store'
	DEDSP = 'Design Example \ DSP'

	for row in range(2, sheet.max_row + 1):
		if sheet['A' + str(row)].value == DE:
			sheet['C' + str(row)] = 'Design examples stored on the Design store that are regression tested.'
		elif sheet['A' + str(row)].value == T:
			sheet['C' + str(row)] = 'These are downloadable tutorials that are regression tested within the design store.'
		elif sheet['A' + str(row)].value == DEODS:
			sheet['C' + str(row)] = 'These are links to places like alterawiki and altera.com for design examples not maintained within the design store.'
		elif sheet['A' + str(row)].value == DELODS:
			sheet['C' + str(row)] = 'These are links to rocketboards designs that are not maintained within the design store.'
		elif sheet['A' + str(row)].value == DEDSP:
			sheet['C' + str(row)] = 'These are links to MATLAB files in the Quartus installation. We don\'t track downloads.' 

	#Sort design data based upon total downloads and store in new worksheet	
	designList = sorted(designList, key=lambda view: view[1], reverse=True)
	try :
		sheet = wb.get_sheet_by_name('DownloadsByDesign')
	except KeyError:
		wb.create_sheet(title = 'DownloadsByDesign')
		sheet = wb.get_sheet_by_name('DownloadsByDesign')
	fontobj = Font(bold=True)
	sheet['A1'] = 'Design'
	sheet['A1'].font = fontobj
	sheet['B1'] = 'Total Downloads'
	sheet['B1'].font = fontobj
	i = 2
	for i in range(len(designList)) :
		sheet['A'+str(i+2)] = designList[i][0] 
		sheet['B'+str(i+2)] = designList[i][1]
		
	#Sort family data based upon total downloads and store in new worksheet	
	familyList = sorted(familyList, key=lambda view: view[1], reverse=True)
	try :
		sheet = wb.get_sheet_by_name('DownloadsByFamily')
	except KeyError:
		wb.create_sheet(title = 'DownloadsByFamily')
		sheet = wb.get_sheet_by_name('DownloadsByFamily')
	fontobj = Font(bold=True)
	sheet['A1'] = 'Family'
	sheet['A1'].font = fontobj
	sheet['B1'] = 'Total Downloads'
	sheet['B1'].font = fontobj
	i = 2
	for i in range(len(familyList)) :
		sheet['A'+str(i+2)] = familyList[i][0] 
		sheet['B'+str(i+2)] = familyList[i][1]

	#Sort dev kit data based upon total downloads and store in new worksheet	
	devkitList = sorted(devkitList, key=lambda view: view[1], reverse=True)
	try :
		sheet = wb.get_sheet_by_name('DownloadsByDevelopmentKit')
	except KeyError:
		wb.create_sheet(title = 'DownloadsByDevelopmentKit')
		sheet = wb.get_sheet_by_name('DownloadsByDevelopmentKit')
	fontobj = Font(bold=True)
	sheet['A1'] = 'Development Kit'
	sheet['A1'].font = fontobj
	sheet['B1'] = 'Total Downloads'
	sheet['B1'].font = fontobj
	i = 2
	for i in range(len(devkitList)) :
		sheet['A'+str(i+2)] = devkitList[i][0] 
		sheet['B'+str(i+2)] = devkitList[i][1]

	#Saving the document	
	print 'Saving Excel Document' 	
	try :
		wb.save(in_file[0])
	except IOError :
		print 'ERROR : Please close the Excel file being read by this script'
		exit()

	print 'Thank You For Your Patronage'